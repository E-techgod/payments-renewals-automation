from __future__ import annotations

from collections.abc import Callable
from io import BytesIO
from typing import Any
from zipfile import BadZipFile

from googleapiclient.discovery import build  # type: ignore[import-untyped]
from googleapiclient.errors import HttpError  # type: ignore[import-untyped]
from httplib2 import HttpLib2Error  # type: ignore[import-untyped]
from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import (  # type: ignore[import-untyped]
    InvalidFileException,
)

from app.config import Config
from app.retry import retry_with_backoff
from app.spreadsheet.base import (
    SpreadsheetError,
    SpreadsheetProvider,
    build_row_dict,
    row_has_client_identity,
    resolve_headers,
)

DriveServiceFactory = Callable[[], Any]
_TRANSIENT_HTTP_STATUSES = frozenset({429, 500, 502, 503, 504})


class _RetryableSpreadsheetError(Exception):
    """Wrap transient API failures so the retry decorator can target them."""


class XlsxDriveProvider(SpreadsheetProvider):
    def __init__(self, config: Config, service_factory: DriveServiceFactory) -> None:
        self._config = config
        self._service_factory = service_factory

    @classmethod
    def from_credentials(cls, config: Config, credentials: Any) -> XlsxDriveProvider:
        def service_factory() -> Any:
            return build("drive", "v3", credentials=credentials)

        return cls(config=config, service_factory=service_factory)

    def load_rows(self) -> list[dict[str, object]]:
        try:
            workbook_bytes = self._download_workbook_bytes_with_retry()
        except (
            _RetryableSpreadsheetError,
            TimeoutError,
            ConnectionError,
        ) as exc:
            raise SpreadsheetError(
                "Drive API request failed after retries were exhausted"
            ) from exc
        try:
            workbook = load_workbook(workbook_bytes, data_only=True, read_only=True)
        except (
            BadZipFile,
            InvalidFileException,
            KeyError,
            ValueError,
        ) as exc:
            raise SpreadsheetError(
                "Downloaded file is not a valid .xlsx workbook"
            ) from exc
        try:
            worksheet = workbook.worksheets[0]
            #print("WORKBOOK SHEETS:", workbook.sheetnames)
            #print("SELECTED SHEET:", workbook.worksheets[0].title)
            row_iterator = worksheet.iter_rows(values_only=True)
            try:
                raw_header_row = next(row_iterator)
            except StopIteration:
                return []

            header_row = ["" if cell is None else str(cell) for cell in raw_header_row]
            resolved_headers = resolve_headers(header_row, self._config)
            rows: list[dict[str, object]] = []
            for raw_row in row_iterator:
                materialized_row = list(raw_row)
                if not row_has_client_identity(
                    materialized_row, resolved_headers, self._config
                ):
                    break
                rows.append(build_row_dict(materialized_row, resolved_headers))

            """For debug
            for idx, row in enumerate(rows[:5]):
                print(f"\n--- Row {idx + 1} ---")
                for key, value in row.items():
                    print(f"  {key}: {value}")"""

            return rows

        finally:
            workbook.close()

    def _download_workbook_bytes_with_retry(self) -> BytesIO:
        @retry_with_backoff(
            max_attempts=self._config.retry_max_attempts,
            base_delay_seconds=self._config.retry_base_delay_seconds,
            retryable_exceptions=(
                _RetryableSpreadsheetError,
                TimeoutError,
                ConnectionError,
            ),
        )
        def download_workbook_bytes() -> BytesIO:
            return self._download_workbook_bytes()

        return download_workbook_bytes()

    def _download_workbook_bytes(self) -> BytesIO:
        service = self._service_factory()
        try:
            request = service.files().get_media(
                fileId=self._config.google_drive_file_id
            )
            workbook_buffer = BytesIO(request.execute())
        except HttpError as exc:
            if exc.resp.status in _TRANSIENT_HTTP_STATUSES:
                raise _RetryableSpreadsheetError("Transient Drive API failure") from exc
            raise SpreadsheetError("Drive API request failed") from exc
        except HttpLib2Error as exc:
            raise _RetryableSpreadsheetError(
                "Transient Drive transport failure"
            ) from exc
        workbook_buffer.seek(0)
        return workbook_buffer
